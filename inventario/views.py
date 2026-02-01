from django.shortcuts import render, redirect, get_object_or_404
from django import forms
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Producto, Movimiento, TipoMovimiento
from django.contrib.auth.signals import user_logged_in
from django.dispatch import receiver
from django.contrib.contenttypes.models import ContentType
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from django.db.models import Q, Sum, Value
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE, DELETION
from django.http import HttpResponse # Importar para exportar
from io import BytesIO # Importar para manejar el buffer de texto para Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import TYPE_CHECKING, Any
from django.utils import timezone

if TYPE_CHECKING:
    from django.contrib.auth.models import User
user_model = get_user_model()
from .forms import ProductoForm

@login_required
def dashboard(request):
    """
    Vista para mostrar el dashboard principal con las tarjetas de resumen.
    Los datos se filtran por usuario si no es un superusuario.
    """
    context = {}
    
    if not request.user.is_superuser:
        # Filtrar por usuario normal
        user_product_ids = Movimiento.objects.filter(usuario=request.user).values_list('producto_id', flat=True).distinct()
        productos_activos = Producto.objects.filter(id__in=user_product_ids, activo=True)
        productos_inactivos = Producto.objects.filter(id__in=user_product_ids, activo=False)
        movimientos_usuario = Movimiento.objects.filter(usuario=request.user)
        
        context['total_bienes_activos'] = productos_activos.count()
        context['total_bienes_inactivos'] = productos_inactivos.count()
        context['total_movimientos'] = movimientos_usuario.count()
        valor_total = productos_activos.aggregate(total=Sum('valor_unitario'))['total']
        context['valor_total_inventario'] = valor_total or 0
        context['total_usuarios'] = user_model.objects.filter(id=request.user.id).count() # Solo se cuenta a sí mismo
        context['movimientos_recientes'] = movimientos_usuario.filter(fecha__gte=timezone.now() - timezone.timedelta(days=7)).count()
    else:
        # El superusuario ve todo
        productos_activos = Producto.objects.filter(activo=True)
        context['total_bienes_activos'] = productos_activos.count()
        context['total_bienes_inactivos'] = Producto.objects.filter(activo=False).count()
        context['total_movimientos'] = Movimiento.objects.count()
        valor_total = productos_activos.aggregate(total=Sum('valor_unitario'))['total']
        context['valor_total_inventario'] = valor_total or 0
        context['total_usuarios'] = user_model.objects.count()
        context['movimientos_recientes'] = Movimiento.objects.filter(fecha__gte=timezone.now() - timezone.timedelta(days=7)).count()

    return render(request, 'dashboard.html', context)

@login_required
def historial_view(request):
    """
    Vista para mostrar el historial completo de movimientos.
    Filtrado por usuario si no es superusuario.
    """
    query = request.GET.get('q', '')
    
    if not request.user.is_superuser:
        movimientos = Movimiento.objects.filter(usuario=request.user).select_related('producto', 'usuario', 'tipo_movimiento').order_by('-fecha')
    else:
        movimientos = Movimiento.objects.all().select_related('producto', 'usuario', 'tipo_movimiento').order_by('-fecha')
    
    if query:
        # Un superusuario puede buscar por cualquier usuario, uno normal solo en sus propios movimientos
        search_base = movimientos
        if not request.user.is_superuser:
            search_base = movimientos.filter(usuario=request.user)

        movimientos = search_base.filter(
            Q(usuario__username__icontains=query) |
            Q(producto__descripcion__icontains=query) |
            Q(observacion__icontains=query)
        )

    context = {
        'movimientos': movimientos,
        'current_query': query,
    }
    return render(request, 'historial.html', context)

@login_required
@require_POST
def limpiar_historial_movimientos(request):
    """
    Vista para eliminar todo el historial de movimientos.
    """
    Movimiento.objects.all().delete()
    return redirect('historial')

@login_required
@require_POST
def limpiar_historial_usuarios(request):
    """
    Vista para eliminar todo el historial de actividad de usuarios (LogEntry).
    """
    LogEntry.objects.all().delete()
    return redirect('usuarios_log')

# Inventario principal
@login_required
def inventario(request):
    # Lógica de Búsqueda
    query = request.GET.get('q', '')
    
    # Lógica de Ordenación
    sort_by = request.GET.get('sort', 'fecha_agregado')
    direction = request.GET.get('dir', 'desc')

    # Si el usuario no es superusuario, filtramos por sus productos
    if not request.user.is_superuser:
        # 1. Obtener los IDs de los productos con los que el usuario ha interactuado
        user_product_ids = Movimiento.objects.filter(usuario=request.user).values_list('producto_id', flat=True).distinct()
        
        # 2. Filtrar los productos basándose en esos IDs
        productos_qs = Producto.objects.filter(id__in=user_product_ids, activo=True)
        
        # 3. Filtrar también los movimientos que se muestran
        movimientos = Movimiento.objects.filter(usuario=request.user).order_by('-fecha')[:20]

    else:
        # El superusuario ve todos los productos y movimientos
        productos_qs = Producto.objects.filter(activo=True)
        movimientos = Movimiento.objects.all().order_by('-fecha')[:20]

    if query:
        productos_qs = productos_qs.filter(
            Q(descripcion__icontains=query) |
            Q(codigo__icontains=query) |
            Q(num_de_bien__icontains=query) |
            Q(cod_bien__icontains=query) |
            Q(concepto__icontains=query)
        )

    # Validar y aplicar ordenación
    allowed_sort_fields = ['cantidad', 'codigo', 'num_de_bien', 'descripcion', 'cod_bien', 'concepto', 'valor_unitario', 'fecha_agregado']
    if sort_by in allowed_sort_fields:
        if direction == 'desc':
            sort_by = f'-{sort_by}'
        productos = productos_qs.order_by(sort_by)
    else:
        productos = productos_qs.order_by('-fecha_agregado') # Orden por defecto

    return render(request, 'index.html', {
        'productos': productos,
        'movimientos': movimientos,
        'current_query': query,
        'current_sort': sort_by.lstrip('-'),
        'current_dir': direction,
    })

# Panel de administración con pestañas (historial + productos ocultos + usuarios)
def _is_staff_user(u: Any) -> bool:
    return getattr(u, 'is_staff', False)

@user_passes_test(_is_staff_user)
def admin_panel(request):
    movimientos = Movimiento.objects.all().order_by('-fecha')
    ocultos = Producto.objects.filter(activo=False)
    
    query = request.GET.get('q', '')
    if query:
        usuarios = user_model.objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    else:
        usuarios = user_model.objects.all()

    return render(request, 'admin_panel.html', {
        'movimientos': movimientos,
        'ocultos': ocultos,
        'usuarios': usuarios,
        'current_query': query,
    })

# Vista para Bienes Inactivos
@login_required
def bienes_inactivos(request):
    query = request.GET.get('q', '')
    
    if not request.user.is_superuser:
        user_product_ids = Movimiento.objects.filter(usuario=request.user).values_list('producto_id', flat=True).distinct()
        ocultos_qs = Producto.objects.filter(id__in=user_product_ids, activo=False).order_by('-deleted_at')
    else:
        ocultos_qs = Producto.objects.filter(activo=False).order_by('-deleted_at')

    if query:
        ocultos_qs = ocultos_qs.filter(
            Q(descripcion__icontains=query) |
            Q(codigo__icontains=query) |
            Q(num_de_bien__icontains=query) |
            Q(cod_bien__icontains=query) |
            Q(concepto__icontains=query)
        )

    return render(request, 'bienesocu.html', {
        'ocultos': ocultos_qs,
        'current_query': query,
    })

# Vista para Bienes Ocultos (similar a bienes_inactivos)
@login_required
def bienes_ocultos(request):
    query = request.GET.get('q', '')
    if not request.user.is_superuser:
        user_product_ids = Movimiento.objects.filter(usuario=request.user).values_list('producto_id', flat=True).distinct()
        ocultos_qs = Producto.objects.filter(id__in=user_product_ids, activo=False).order_by('-deleted_at')
    else:
        ocultos_qs = Producto.objects.filter(activo=False).order_by('-deleted_at')

    if query:
        ocultos_qs = ocultos_qs.filter(
            Q(descripcion__icontains=query) |
            Q(codigo__icontains=query) |
            Q(num_de_bien__icontains=query) |
            Q(cod_bien__icontains=query) |
            Q(concepto__icontains=query)
        )

    return render(request, 'bienesocu.html', {
        'ocultos': ocultos_qs,
        'current_query': query,
    })

# Vista para exportar productos a Excel
@login_required
def export_productos_excel(request):
    if not request.user.is_superuser:
        user_product_ids = Movimiento.objects.filter(usuario=request.user).values_list('producto_id', flat=True).distinct()
        productos = Producto.objects.filter(id__in=user_product_ids, activo=True).order_by('descripcion')
    else:
        productos = Producto.objects.filter(activo=True).order_by('descripcion')

    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos de Inventario"

    # Definir estilos
    header_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Encabezados
    headers = ["Cantidad", "CODIGO", "NumDeBien", "Descripción", "CodBien", "Concepto", "Valor unitario", "Fecha agregado"]
    ws.append(headers)

    # Aplicar estilos a los encabezados
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = thin_border
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 20 # Ancho por defecto

    # Datos
    for producto in productos:
        row_data = [
            producto.cantidad,
            producto.codigo,
            producto.num_de_bien,
            producto.descripcion,
            producto.cod_bien,
            producto.concepto,
            float(producto.valor_unitario), # Asegurarse que es un float para Excel
            producto.fecha_agregado.strftime("%Y-%m-%d %H:%M:%S") if producto.fecha_agregado else ""
        ]
        ws.append(row_data)
        # Aplicar bordes a las celdas de datos
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Ajustar ancho de columnas al contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Obtener el nombre de la columna (e.g., 'A')
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el libro de trabajo en un buffer de memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(excel_file.getvalue(), 
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productos_inventario.xlsx"'
    return response


# Vista para importar productos desde Excel
@login_required
@require_POST
def importar_productos_excel(request):
    from django.contrib import messages
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Validar que es un archivo Excel
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo no es un Excel válido (.xlsx o .xls).')
            return redirect('inventario')

        try:
            # Cargar el libro de trabajo
            wb = load_workbook(excel_file)
            ws = wb.active # Seleccionar la hoja activa

            productos_creados = 0
            productos_actualizados = 0

            # Asumiendo que la primera fila son los encabezados
            headers = [cell.value for cell in ws[1]]
            
            # Mapear encabezados a campos del modelo (simple, puedes personalizar esto)
            # Asegurarse de que los encabezados del Excel coincidan con esto
            header_mapping = {
                "Cantidad": "cantidad",
                "CODIGO": "codigo",
                "NumDeBien": "num_de_bien",
                "Descripción": "descripcion",
                "CodBien": "cod_bien",
                "Concepto": "concepto",
                "Valor unitario": "valor_unitario",
            }
            
            # Obtener índices de las columnas
            col_indices = {header_mapping[h]: headers.index(h) for h in headers if h in header_mapping}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2): # Saltar encabezado
                row_data = {}
                for field_name, col_idx in col_indices.items():
                    row_data[field_name] = row[col_idx]

                try:
                    codigo = row_data.get('codigo')
                    num_de_bien = row_data.get('num_de_bien')
                    
                    if not codigo: # Asegurar que haya al menos un identificador
                        messages.warning(request, f"Fila {row_idx}: Ignorada. Falta 'CODIGO'.")
                        continue

                    producto, created = Producto.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'descripcion': row_data.get('descripcion'),
                            'num_de_bien': num_de_bien,
                            'cod_bien': row_data.get('cod_bien'),
                            'concepto': row_data.get('concepto'),
                            'cantidad': int(row_data.get('cantidad', 0)),
                            'valor_unitario': float(row_data.get('valor_unitario', 0.0)),
                            'activo': True 
                        }
                    )
                    
                    if created:
                        productos_creados += 1
                        tipo_entrada, _ = TipoMovimiento.objects.get_or_create(tipo='entrada')
                        Movimiento.objects.create(
                            producto=producto,
                            tipo_movimiento=tipo_entrada,
                            cantidad=producto.cantidad,
                            usuario=request.user,
                            observacion=f"Producto importado vía Excel (nuevo). Cantidad: {producto.cantidad}"
                        )
                    else:
                        productos_actualizados += 1
                        tipo_edicion, _ = TipoMovimiento.objects.get_or_create(tipo='edicion')
                        Movimiento.objects.create(
                            producto=producto,
                            tipo_movimiento=tipo_edicion,
                            cantidad=producto.cantidad,
                            usuario=request.user,
                            observacion=f"Producto importado vía Excel (actualizado). Cantidad en Excel: {row_data.get('cantidad', 0)}"
                        )

                except Exception as e:
                    messages.error(request, f"Error al procesar fila {row_idx} ({row_data}): {e}")
                    continue
            
            messages.success(request, f'Importación Excel completada: {productos_creados} creados, {productos_actualizados} actualizados.')

        except Exception as e:
            messages.error(request, f"Error al leer el archivo Excel: {e}")
            
    return redirect('inventario')

# Agregar producto
@login_required
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            if not getattr(producto, 'nombre', None):
                producto.nombre = (producto.descripcion or '')[:100]
            producto.save()
            tipo_entrada, _ = TipoMovimiento.objects.get_or_create(tipo='entrada')
            Movimiento.objects.create(
                producto=producto,
                tipo_movimiento=tipo_entrada,
                cantidad=producto.cantidad,
                usuario=request.user,
                observacion="Producto agregado"
            )
            return redirect('inventario')
    else:
        form = ProductoForm()
    return render(request, 'agregar_producto.html', {'form': form})

# Editar producto
@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            tipo_edicion, _ = TipoMovimiento.objects.get_or_create(tipo='edicion')
            Movimiento.objects.create(
                producto=producto,
                tipo_movimiento=tipo_edicion,
                cantidad=producto.cantidad,
                usuario=request.user,
                observacion="Producto editado"
            )
            return redirect('inventario')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'editar_producto.html', {'form': form, 'producto': producto})

# Ocultar producto
@login_required
@require_POST
def ocultar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    tipo_ocultar, _ = TipoMovimiento.objects.get_or_create(tipo='INACTIVO')
    Movimiento.objects.create(
        producto=producto,
        tipo_movimiento=tipo_ocultar,
        cantidad=producto.cantidad,
        usuario=request.user,
        observacion="Producto marcado como inactivo"
    )
    producto.delete()
    return redirect('inventario')

# Restaurar producto
@login_required
@require_POST
def restaurar_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id, activo=False)
    producto.restore()
    tipo_restaurar, _ = TipoMovimiento.objects.get_or_create(tipo='restaurar')
    Movimiento.objects.create(
        producto=producto,
        tipo_movimiento=tipo_restaurar,
        cantidad=producto.cantidad,
        usuario=request.user,
        observacion="Producto marcado como activo"
    )
    return redirect(request.META.get('HTTP_REFERER', 'inventario'))

# Activar/Desactivar usuario
@login_required
@require_POST
@user_passes_test(lambda u: u.is_superuser)
def toggle_user_active(request, user_id):
    user_to_toggle = get_object_or_404(user_model, id=user_id)
    if user_to_toggle != request.user: 
        user_to_toggle.is_active = not user_to_toggle.is_active
        user_to_toggle.save()
        action_message = "Usuario activado." if user_to_toggle.is_active else "Usuario desactivado."
        LogEntry.objects.create(
            user_id=request.user.pk,
            content_type_id=ContentType.objects.get_for_model(user_to_toggle).pk,
            object_id=user_to_toggle.pk,
            object_repr=str(user_to_toggle),
            action_flag=2,
            change_message=action_message
        )
    return redirect('admin_panel')

# Dar/Quitar permisos de administrador
@login_required
@require_POST
@user_passes_test(lambda u: u.is_superuser)
def toggle_user_staff(request, user_id):
    user_to_toggle = get_object_or_404(user_model, id=user_id)
    if user_to_toggle != request.user:
        user_to_toggle.is_staff = not user_to_toggle.is_staff
        user_to_toggle.save()
        action_message = "Se concedieron permisos de administrador." if user_to_toggle.is_staff else "Se revocaron permisos de administrador."
        LogEntry.objects.create(
            user_id=request.user.pk,
            content_type_id=ContentType.objects.get_for_model(user_to_toggle).pk,
            object_id=user_to_toggle.pk,
            object_repr=str(user_to_toggle),
            action_flag=2,
            change_message=action_message
        )
    return redirect('admin_panel')

# --- INICIO: Registrar inicios de sesión ---
@receiver(user_logged_in)
def log_user_login(sender, request, user, **kwargs):
    LogEntry.objects.create(
        user_id=user.pk,
        content_type_id=ContentType.objects.get_for_model(user).pk,
        object_id=user.pk,
        object_repr=str(user),
        action_flag=2,
        change_message='Inició sesión.',
    )
# --- FIN: Registrar inicios de sesión ---

# Log de actividad de usuarios
@login_required
def user_log_view(request):
    if not request.user.is_superuser:
        log_entries = LogEntry.objects.filter(user_id=request.user.id).select_related('user', 'content_type').order_by('-action_time')
    else:
        log_entries = LogEntry.objects.select_related('user', 'content_type').order_by('-action_time')

    query = request.GET.get('q', '')
    if query:
        search_filters = (
            Q(user__username__icontains=query) |
            Q(object_repr__icontains=query) |
            Q(change_message__icontains=query)
        )
        log_entries = log_entries.filter(search_filters)
        
    return render(request, 'usuarioslog.html', {'log_entries': log_entries})
